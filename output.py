import openpyxl


def write(store, date, items, file):
    #date = date.strftime('%d.%m.%Y %H:%M')
    wb = openpyxl.load_workbook(file)
    page = wb.worksheets[0]
    for max_row, row in enumerate(page, 1):
        if all(c.value is None for c in row):
            break
    if any(i.value == date for i in page['A']):
        return
    for i, item in enumerate(items, max_row):
        page.cell(i, 1, date.strftime('%d.%m.%Y %H:%M'))
        page.cell(i, 2, item['name'])
        page.cell(i, 3, item['price'])
        page.cell(i, 4, item['quantity'])
        page.cell(i, 5, item['price']*item['quantity'])
        page.cell(i, 6, store)
    wb.save(file)